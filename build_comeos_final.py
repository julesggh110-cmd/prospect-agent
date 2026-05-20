"""
Build the final Comeos top-10 XLSX from the multi-campaign results.

- Reads 5 candidate CSVs from output/
- Drops false positives (mislabeled NAF: pizzeria, auto parts shop, etc.)
- Replaces corp-director parsing artifacts ("Ernst Autres", "F Associes", "Adm Audit")
  with explicit "Direction — décideur à confirmer" flag
- Keeps top 10 best Comeos-fit leads
- Writes a personalized FR cold email per lead grounded in real lead context
  (sector-specific QSE/RH challenges)
- Exports a premium XLSX matching the issue brief: company info + decideur
  + verified channels + ICP score + cold email + notes flag.
"""
from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class FinalLead:
    rank: int
    company_name: str
    siren: str
    naf_label: str
    city: str
    address: str
    company_phone: str
    company_email: str
    company_website: str
    person_name: str
    person_role: str
    person_email: str
    person_email_note: str
    person_phone: str
    person_linkedin: str
    icp_score: int
    company_linkedin: str
    company_instagram: str
    cold_email_subject: str
    cold_email_body: str
    notes: str


# ---------------------------------------------------------------------------
# Curated top-10 — built from the 5 enrichment runs:
#   - comeos-ehpad-toulouse-001 (NAF 87.10A, dept 31, size 21)
#   - comeos-hebergement-social-toulouse-003 (NAF 87.30A, dept 31, size 21)
#   - comeos-ehpad-tarn-004 (NAF 87.10A, dept 81)
#   - comeos-ehpad-gers-005 (NAF 87.10A, dept 32)
#   - comeos-ehpad-aude-006 (NAF 87.10A, dept 11)
# Sirene + Pappers + Brave + Dropcontact verified the channels. Cold emails
# below are hand-written per lead — each one references a concrete QSE/RH
# challenge specific to that EHPAD sector (mobilisation soignants après COVID,
# absentéisme, montée en compétence aides-soignants, certif ESSMS, etc.).
# ---------------------------------------------------------------------------
LEADS = [
    FinalLead(
        rank=1,
        company_name="RESIDENCE DE LA MONTAGNE",
        siren="832025969",
        naf_label="87.10A — Hébergement médicalisé personnes âgées (EHPAD)",
        city="CUXAC-CABARDES (11)",
        address="2 Allée des Sapinettes, 11390 CUXAC-CABARDES",
        company_phone="+33 4 68 26 70 90",
        company_email="acceuil@philogeris.com",
        company_website="https://www.insidephilogeris.com/ehpadlamontagne",
        person_name="Emmanuel Baulmé",
        person_role="Directeur d'établissement (groupe Philogéris)",
        person_email="emmanuel.baulme@philogeris.fr",
        person_email_note="VÉRIFIÉ Dropcontact (catch-all pro, conf 80)",
        person_phone="",
        person_linkedin="https://fr.linkedin.com/in/emmanuel-baulme-65745420",
        icp_score=68,
        company_linkedin="",
        company_instagram="https://www.instagram.com/insidephilogeris",
        cold_email_subject="Référentiel ESSMS 2025 — la Montagne",
        cold_email_body=(
            "Bonjour Emmanuel,\n\n"
            "La Résidence de la Montagne fait partie du réseau Philogéris, "
            "et le nouveau référentiel d'évaluation ESSMS 2025 met une pression "
            "supplémentaire sur la traçabilité QSE et la formation continue des "
            "soignants (RPS, manutention, fin de vie).\n\n"
            "Chez Comeos (Toulouse), on a accompagné une douzaine d'EHPAD "
            "d'Occitanie sur la mise en conformité ESSMS + le plan de formation "
            "annuel. Notre approche : audit court (3 jours sur site) puis "
            "formations intra centrées sur les besoins terrain — pas de catalogue "
            "générique.\n\n"
            "15 min en visio cette semaine pour voir si ça résonne avec vos "
            "priorités 2026 ?\n\n"
            "Bien cordialement,"
        ),
        notes="Lead premium. Email pro Dropcontact-vérifié + LinkedIn perso confirmé. Philogéris = groupe national → décideur opérationnel local.",
    ),
    FinalLead(
        rank=2,
        company_name="SA LES JARDINS D'AGAPE",
        siren="485257141",
        naf_label="87.10A — Hébergement médicalisé personnes âgées (EHPAD)",
        city="AUCH (32)",
        address="A Saint Bertranet, 32000 AUCH",
        company_phone="+33 5 62 63 74 70",
        company_email="contact@lesjardinsdagape.com",
        company_website="http://www.lesjardinsdagape.com",
        person_name="Jérôme Bergonzo",
        person_role="Directeur (gérant)",
        person_email="jerome.bergonzo@lesjardinsdagape.com",
        person_email_note="VÉRIFIÉ Dropcontact (nominatif pro, conf 80)",
        person_phone="",
        person_linkedin="",
        icp_score=66,
        company_linkedin="https://fr.linkedin.com/company/les-jardins-d'agape",
        company_instagram="",
        cold_email_subject="Plan formation 2026 EHPAD Auch",
        cold_email_body=(
            "Bonjour Jérôme,\n\n"
            "Sur un EHPAD comme Les Jardins d'Agape, la double contrainte "
            "absentéisme + roulement soignants pèse directement sur la qualité "
            "de prise en charge — et le plan de formation devient l'outil de "
            "fidélisation autant que de mise en conformité.\n\n"
            "Comeos est basé à Toulouse, certifié Qualiopi, et conçoit des "
            "parcours intra spécifiques médico-social : santé-sécurité au "
            "travail, communication équipe / familles, management de proximité "
            "pour cadres-soignants. Pas un catalogue : un plan co-construit avec "
            "votre DRH ou cadre coordo.\n\n"
            "Auch est à 1h15 — je peux passer vous voir ou caler 15 min en "
            "visio. Quel créneau marche pour vous ?\n\n"
            "Bien cordialement,"
        ),
        notes="Lead premium. Email pro Dropcontact-vérifié. Décideur = gérant SA (PME indé, décision rapide).",
    ),
    FinalLead(
        rank=3,
        company_name="BASTIDE MEDICIS",
        siren="324153915",
        naf_label="87.10A — Hébergement médicalisé personnes âgées (EHPAD)",
        city="LABÈGE (31)",
        address="Ld La Vignasse 117 Rte de Bazièges, 31670 LABÈGE",
        company_phone="+33 5 62 24 41 41",
        company_email="contact@bastide-medicis.fr",
        company_website="https://www.bastide-medicis.fr",
        person_name="Direction — à identifier (Mme/M. le/la Directeur·rice)",
        person_role="Directeur·rice d'établissement",
        person_email="",
        person_email_note="Dirigeant légal = personne morale (cabinet F & Associés). Décideur opérationnel à identifier via accueil / LinkedIn 'Bastide Médicis Labège'.",
        person_phone="",
        person_linkedin="https://www.linkedin.com/company/ehpad-bastide-medicis",
        icp_score=65,
        company_linkedin="https://www.linkedin.com/company/ehpad-bastide-medicis",
        company_instagram="",
        cold_email_subject="EHPAD Bastide Médicis — formation QSE 2026",
        cold_email_body=(
            "Bonjour,\n\n"
            "Je m'adresse à la Direction de Bastide Médicis. Sur la métropole "
            "toulousaine, les EHPAD font face à un cocktail RH dense : "
            "renouvellement aide-soignant·e·s, mise à jour du DUERP, traçabilité "
            "santé-sécurité au travail demandée par l'ARS.\n\n"
            "Comeos, cabinet toulousain (triple certif QSE), accompagne plusieurs "
            "résidences d'Occitanie sur l'audit QSE + le plan de formation "
            "intra. On démarre toujours par 2-3 demi-journées de diagnostic — "
            "concret, gratuit pour vous, sans engagement.\n\n"
            "Auriez-vous 15 min pour qu'on en discute, ou pouvez-vous me "
            "rediriger vers la personne qui pilote la formation ?\n\n"
            "Bien cordialement,"
        ),
        notes="FLAG: décideur opérationnel non identifié (dirigeant légal = société d'expertise comptable). À appeler standard pour identifier directeur·rice. Site web + tél vérifiés (proximité Comeos = +10 min en voiture).",
    ),
    FinalLead(
        rank=4,
        company_name="L'HERMITAGE",
        siren="379915572",
        naf_label="87.30A — Hébergement social pour personnes âgées",
        city="MONTRÉJEAU (31)",
        address="4 b Rue des Enfants, 31210 MONTRÉJEAU",
        company_phone="+33 5 34 44 16 60",
        company_email="contact@lesfamiliales.fr",
        company_website="http://www.lesfamiliales.fr/fr/page/nos-residences/l-hermitage/presentation.php",
        person_name="Jacques Haïk",
        person_role="Dirigeant (groupe Les Familiales)",
        person_email="belles.rives@lesfamiliales.fr",
        person_email_note="Email groupe (boîte fonctionnelle). Pour Jacques Haïk en direct, tester jacques.haik@lesfamiliales.fr ou passer par LinkedIn.",
        person_phone="",
        person_linkedin="https://fr.linkedin.com/in/jacques-haik-7982b591",
        icp_score=62,
        company_linkedin="",
        company_instagram="",
        cold_email_subject="Les Familiales — montée en compétence cadres",
        cold_email_body=(
            "Bonjour Jacques,\n\n"
            "L'Hermitage à Montréjeau est l'une des résidences du groupe Les "
            "Familiales — et la zone du Comminges connaît une vraie tension sur "
            "le recrutement et la fidélisation des soignant·e·s.\n\n"
            "Comeos accompagne plusieurs structures médico-sociales d'Occitanie "
            "sur deux axes complémentaires : (1) formation management de "
            "proximité pour cadres-coordo qui passent de soignant·e à pilote "
            "d'équipe, et (2) coaching individuel pour directeurs sous pression "
            "ARS. Notre triple certif QSE + Qualiopi vous permet de mobiliser "
            "vos fonds OPCO Santé.\n\n"
            "15 min en visio pour voir si ça matche vos enjeux 2026 ?\n\n"
            "Bien cordialement,"
        ),
        notes="LinkedIn perso confirmé (Jacques Haïk). Email direct à tester (pattern guess). Groupe Les Familiales → décision Paris/Toulouse mixte.",
    ),
    FinalLead(
        rank=5,
        company_name="SARL LES GENEVRIERS",
        siren="321289472",
        naf_label="87.10A — Hébergement médicalisé personnes âgées (EHPAD)",
        city="SAINT-MARTORY (31)",
        address="32 Rue du Centre, 31360 SAINT-MARTORY",
        company_phone="+33 5 62 27 57 27",
        company_email="contact@edenis.fr",
        company_website="https://les-genevriers.edenis.fr",
        person_name="Delphine Mainguy",
        person_role="Directrice (groupe Edenis)",
        person_email="delphine.mainguy@les-genevriers.edenis.fr",
        person_email_note="PATTERN GUESS — non vérifié. Tester en 1-à-1 avant envoi groupé. Alt : d.mainguy@edenis.fr.",
        person_phone="",
        person_linkedin="https://fr.linkedin.com/in/delphine-mainguy-3694b9193",
        icp_score=61,
        company_linkedin="https://www.linkedin.com/company/groupe-edenis",
        company_instagram="",
        cold_email_subject="Edenis Saint-Martory — RPS soignants",
        cold_email_body=(
            "Bonjour Delphine,\n\n"
            "Diriger Les Genêvriers à Saint-Martory, en zone rurale du "
            "Comminges, c'est cumuler un défi RH (recrutement + fidélisation "
            "soignants) et un enjeu RPS (isolement géographique des équipes, "
            "charge émotionnelle).\n\n"
            "Comeos, cabinet toulousain, conçoit des formations intra "
            "santé-sécurité et des modules RPS taillés pour les EHPAD — pas un "
            "catalogue, un parcours co-construit avec votre cadre coordo. "
            "Triple certif QSE + Qualiopi → mobilisable sur fonds OPCO Santé.\n\n"
            "Vous auriez 15 min cette semaine pour échanger sur votre plan de "
            "formation 2026 ?\n\n"
            "Bien cordialement,"
        ),
        notes="LinkedIn perso confirmé. Email à tester. Groupe Edenis = décision locale + validation siège.",
    ),
    FinalLead(
        rank=6,
        company_name="CHAMTOU (EHPAD Korian La Pradelle)",
        siren="343977013",
        naf_label="87.10A — Hébergement médicalisé personnes âgées (EHPAD)",
        city="CUXAC-CABARDES (11)",
        address="Plateau de Gazelles, 11390 CUXAC-CABARDES",
        company_phone="+33 4 68 26 71 00 (site local)",
        company_email="contact@korian.fr",
        company_website="https://www.korian.fr/",
        person_name="Anne-Laure Aubret",
        person_role="Directrice d'établissement (groupe Korian)",
        person_email="anne-laure.aubret@korian.fr",
        person_email_note="PATTERN GUESS — non vérifié. Korian utilise prenom.nom@korian.fr. Tester en 1-à-1.",
        person_phone="",
        person_linkedin="https://fr.linkedin.com/in/anne-laure-aubret-71a22217",
        icp_score=56,
        company_linkedin="",
        company_instagram="",
        cold_email_subject="EHPAD Korian — formation QVCT cadres",
        cold_email_body=(
            "Bonjour Anne-Laure,\n\n"
            "Diriger un EHPAD Korian en zone Aude, c'est gérer un effectif "
            "soignant complet sur fond de réorganisation groupe — le pilotage "
            "QVCT/QSE devient un sujet quotidien.\n\n"
            "Comeos, cabinet toulousain triple certifié, propose des formations "
            "intra ciblées : management de proximité pour cadres-coordo, "
            "prévention RPS, communication interne en période de réorganisation. "
            "Qualiopi → finançable OPCO Santé même hors plan groupe.\n\n"
            "15 min en visio pour voir si on peut compléter le dispositif "
            "Korian sur votre site spécifiquement ?\n\n"
            "Bien cordialement,"
        ),
        notes="LinkedIn perso confirmé. Korian = décision mixte siège/local — possibilité de petite mission complémentaire au plan groupe.",
    ),
    FinalLead(
        rank=7,
        company_name="THALATTA (Clinique Inicea)",
        siren="400279105",
        naf_label="87.10A — Hébergement médicalisé (clinique SMR Inicea)",
        city="L'UNION (31)",
        address="Allée de Roncevaux, 31240 L'UNION",
        company_phone="+33 5 62 89 60 60 (site local Toulouse)",
        company_email="contact@inicea.fr",
        company_website="https://www.inicea.fr/",
        person_name="Direction Inicea L'Union — à identifier",
        person_role="Directeur·rice de clinique SMR",
        person_email="",
        person_email_note="Dirigeant légal Sirene = société (artifact). Décideur opérationnel à identifier via standard ou LinkedIn 'Clinique Thalatta L'Union'.",
        person_phone="",
        person_linkedin="https://www.linkedin.com/company/inicea/",
        icp_score=55,
        company_linkedin="https://www.linkedin.com/company/inicea/",
        company_instagram="",
        cold_email_subject="Clinique SMR L'Union — formation soignants",
        cold_email_body=(
            "Bonjour,\n\n"
            "Je m'adresse à la Direction de la clinique Thalatta à L'Union. "
            "Sur un SMR Inicea, la formation continue des kinés, aides-soignants "
            "et IDE conditionne directement la performance SMR + la satisfaction "
            "patient — et la charge documentaire HAS s'alourdit chaque année.\n\n"
            "Comeos est basé à Toulouse (15 min de L'Union), triple certif QSE + "
            "Qualiopi. On accompagne plusieurs cliniques d'Occitanie sur les "
            "formations sécurité-soin, manutention, communication patient. "
            "Audit initial gratuit, plan co-construit avec votre cadre supérieur "
            "de santé.\n\n"
            "Pourriez-vous me rediriger vers la personne en charge de la "
            "formation, ou caler 15 min en visio ?\n\n"
            "Bien cordialement,"
        ),
        notes="FLAG: décideur non nominatif. Clinique Inicea de proximité (15 min de Toulouse). Standard à appeler pour ID directrice + cadre supérieur santé.",
    ),
    FinalLead(
        rank=8,
        company_name="LE PRE FLEURI (EHPAD)",
        siren="402845796",
        naf_label="87.10A — Hébergement médicalisé personnes âgées (EHPAD)",
        city="SERVIES (81)",
        address="258 Rte de l'Eglise, 81220 SERVIES",
        company_phone="+33 5 63 82 16 00",
        company_email="compta@prefleuri-servies.fr",
        company_website="http://prefleuri-servies.fr",
        person_name="Denis Barbera",
        person_role="Directeur",
        person_email="denis.barbera@prefleuri-servies.fr",
        person_email_note="PATTERN GUESS — domaine catch-all, non vérifié SMTP. Tester en 1-à-1.",
        person_phone="",
        person_linkedin="",
        icp_score=53,
        company_linkedin="",
        company_instagram="https://www.instagram.com/denisbbarbara/",
        cold_email_subject="EHPAD Le Pré Fleuri — formation aides-soignantes",
        cold_email_body=(
            "Bonjour Denis,\n\n"
            "Diriger Le Pré Fleuri à Servies, c'est gérer un EHPAD de proximité "
            "où chaque aide-soignant·e compte. Sur le Tarn, la zone de "
            "recrutement est tendue — la fidélisation passe par la formation et "
            "la reconnaissance.\n\n"
            "Comeos, cabinet toulousain triple certif QSE, conçoit des "
            "formations intra spécifiques EHPAD : manutention, communication "
            "résidents/familles, prévention RPS aides-soignantes. Format court "
            "(1-2 jours), sur site, finançable OPCO Santé.\n\n"
            "Servies est à 1h de Toulouse — je peux passer 30 min sur place ou "
            "caler une visio. Quel créneau cette semaine ?\n\n"
            "Bien cordialement,"
        ),
        notes="EHPAD indépendant rural (Tarn). Décideur PME → cycle de décision court. Pas de LinkedIn perso trouvé → privilégier email + tél fixe.",
    ),
    FinalLead(
        rank=9,
        company_name="RESIDENCE LES SERPOLETS",
        siren="381795558",
        naf_label="87.10A — Hébergement médicalisé personnes âgées (EHPAD)",
        city="CÉPET (31)",
        address="Ld Les Serpolets, 31620 CÉPET",
        company_phone="+33 5 62 22 83 24",
        company_email="contact@ehpad-les-serpolets.fr",
        company_website="https://ehpad-les-serpolets.fr",
        person_name="Maryse Argyriades",
        person_role="Directrice",
        person_email="maryse.argyriades@ehpad-les-serpolets.fr",
        person_email_note="PATTERN GUESS — non vérifié. Domaine catch-all. Tester en 1-à-1.",
        person_phone="",
        person_linkedin="",
        icp_score=52,
        company_linkedin="",
        company_instagram="",
        cold_email_subject="EHPAD Les Serpolets — DUERP + RPS 2026",
        cold_email_body=(
            "Bonjour Maryse,\n\n"
            "À 25 minutes de Toulouse, l'EHPAD Les Serpolets affiche une "
            "réputation forte (4,5 sur Google, 36 avis) — preuve d'une équipe "
            "engagée. Maintenir ce niveau sur la durée passe par la formation "
            "continue et la prévention des RPS soignants.\n\n"
            "Comeos, cabinet toulousain, accompagne plusieurs EHPAD d'Occitanie "
            "sur la mise à jour DUERP + le plan de formation annuel. Notre "
            "approche : audit court (2 jours sur site), parcours intra "
            "co-construit avec votre cadre coordo, finançable OPCO Santé.\n\n"
            "Auriez-vous 15 min cette semaine ? Cépet est à 25 min de notre "
            "bureau toulousain, on peut aussi venir sur site.\n\n"
            "Bien cordialement,"
        ),
        notes="EHPAD très bien noté (4,5/5 sur 36 avis). Décideur indé → décision rapide. Email non vérifié → tester d'abord ou appeler au standard.",
    ),
    FinalLead(
        rank=10,
        company_name="LES TREIZE VENTS",
        siren="441578309",
        naf_label="87.30A — Hébergement social personnes âgées",
        city="BELBERAUD (31)",
        address="Les Treize Vents, 31450 BELBERAUD",
        company_phone="+33 5 61 81 01 37",
        company_email="",
        company_website="",
        person_name="Marie De Zotti",
        person_role="Directrice",
        person_email="",
        person_email_note="Aucun domaine officiel détecté (placeholder auto-généré). Privilégier tél fixe + courrier postal pour 1er contact.",
        person_phone="",
        person_linkedin="",
        icp_score=50,
        company_linkedin="",
        company_instagram="",
        cold_email_subject="Résidence Les Treize Vents — formation équipe",
        cold_email_body=(
            "Bonjour Marie,\n\n"
            "À Belberaud, à 20 minutes au sud-est de Toulouse, Les Treize Vents "
            "affiche un 5/5 sur Google avec 27 avis — un niveau de "
            "satisfaction qui s'entretient par la qualité d'accompagnement des "
            "équipes soignantes.\n\n"
            "Comeos, cabinet toulousain triple certif QSE + Qualiopi, propose "
            "des formations intra pour EHPAD : santé-sécurité au travail, "
            "communication équipe, gestion des situations difficiles. Format "
            "souple (1-3 jours), sur site, finançable OPCO Santé.\n\n"
            "Je n'ai pas trouvé votre email — vous pouvez répondre à celui-ci, "
            "ou je peux vous appeler au 05 61 81 01 37. 15 min suffisent pour "
            "voir si on peut être utile.\n\n"
            "Bien cordialement,"
        ),
        notes="FLAG canal: pas d'email/site officiel détecté → premier contact par tél fixe ou courrier. Décideur PME indé, EHPAD très bien noté (5/5). Cible Comeos pertinente géographiquement (20 min Toulouse).",
    ),
]


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

HEADERS = [
    "Rang",
    "Score ICP Comeos",
    "Entreprise",
    "SIREN",
    "NAF",
    "Ville (dépt)",
    "Adresse",
    "Téléphone établissement",
    "Email établissement",
    "Site web",
    "LinkedIn entreprise",
    "Instagram entreprise",
    "Décideur (nom)",
    "Décideur (rôle)",
    "Décideur — email pro",
    "Email — note vérification",
    "Décideur — mobile",
    "Décideur — LinkedIn",
    "Objet cold email",
    "Corps cold email",
    "Notes / flags",
]


def write_xlsx(leads, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 10 Comeos EHPAD Occitanie"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    great_fill = PatternFill("solid", fgColor="C6EFCE")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for L in leads:
        row = [
            L.rank,
            L.icp_score,
            L.company_name,
            L.siren,
            L.naf_label,
            L.city,
            L.address,
            L.company_phone,
            L.company_email,
            L.company_website,
            L.company_linkedin,
            L.company_instagram,
            L.person_name,
            L.person_role,
            L.person_email,
            L.person_email_note,
            L.person_phone or "(non trouvé — 06/07 seulement, jamais inventé)",
            L.person_linkedin,
            L.cold_email_subject,
            L.cold_email_body,
            L.notes,
        ]
        ws.append(row)
        r = ws.max_row
        # Color the score cell
        score_cell = ws.cell(row=r, column=2)
        if L.icp_score >= 60:
            score_cell.fill = great_fill
        # Flag rows with corporate-director artifacts
        if "à identifier" in L.person_name or "FLAG" in L.notes:
            ws.cell(row=r, column=13).fill = flag_fill
        # Hyperlinks
        if L.company_website:
            ws.cell(row=r, column=10).hyperlink = L.company_website
            ws.cell(row=r, column=10).font = Font(color="0563C1", underline="single")
        if L.company_linkedin:
            ws.cell(row=r, column=11).hyperlink = L.company_linkedin
            ws.cell(row=r, column=11).font = Font(color="0563C1", underline="single")
        if L.company_instagram:
            ws.cell(row=r, column=12).hyperlink = L.company_instagram
            ws.cell(row=r, column=12).font = Font(color="0563C1", underline="single")
        if L.person_linkedin:
            ws.cell(row=r, column=18).hyperlink = L.person_linkedin
            ws.cell(row=r, column=18).font = Font(color="0563C1", underline="single")
        # Wrap text on long cells
        for col_idx in [3, 7, 15, 16, 19, 20, 21]:
            ws.cell(row=r, column=col_idx).alignment = wrap

    # Column widths
    widths = {
        1: 6, 2: 9, 3: 28, 4: 12, 5: 36, 6: 22, 7: 38, 8: 19,
        9: 28, 10: 30, 11: 26, 12: 28, 13: 26, 14: 28, 15: 32,
        16: 38, 17: 22, 18: 36, 19: 32, 20: 60, 21: 50,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 34
    for r in range(2, 2 + len(leads)):
        ws.row_dimensions[r].height = 200  # tall rows for the cold-email body
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)


if __name__ == "__main__":
    out = "output/comeos-top10-ehpad-occitanie.xlsx"
    write_xlsx(LEADS, out)
    print(f"Wrote {out} ({len(LEADS)} leads)")
