"""Build the final Comeos IA Formation deliverable (GROUPE A services Occitanie 50-249).

Consolidates 3 sub-campaigns (70.22Z conseil, 78.10Z placement, 82.99Z autres
services support), dedupes by SIREN, backfills NAF labels per source, removes
obviously bogus contact data (MOBIVIA support@1and1.fr → 1and1 hosting catchall),
and writes top-20 by overall_score with per-lead authored FR cold emails.
"""
from __future__ import annotations

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
OUT = ROOT / "output" / "comeos-ia-services-toulouse-FINAL.xlsx"

SOURCES = [
    (ROOT / "output/comeos-ia-services-toulouse.csv",         "70.22Z", "Conseil pour les affaires et autres conseils de gestion"),
    (ROOT / "output/comeos-ia-services-toulouse-cascade1.csv","78.10Z", "Activités des agences de placement de main-d'œuvre"),
    (ROOT / "output/comeos-ia-services-toulouse-cascade2.csv","82.99Z", "Autres activités de soutien aux entreprises n.c.a."),
]

# SIRENs to scrub (manually identified)
DROP_PERSON_EMAIL = {
    "470501545": "MOBIVIA – support@1and1.fr est un catchall hébergeur, pas l'email DRH (CROSS-COMPANY)",
}
# Phone numbers mis-associated to company (geo mismatch)
PHONE_NOTES = {
    "470501545": "Tel +33 7 69 06 79 58 récupéré via OSM ≠ ligne corporate vérifiée → à confirmer",
    "451814644": "Tel +33 1 53 86 84 01 = Paris (01) alors qu'établissement Toulouse 31 → ligne HQ groupe, pas la filiale",
}
WEBSITE_NOTES = {
    "470501545": "Site 'mecaexpertsam31.fr' ne correspond pas à MOBIVIA (catchall OSM)",
}


def load_all_rows() -> list[dict]:
    rows = []
    for path, naf, naf_label in SOURCES:
        with path.open(encoding="utf-8-sig") as f:
            for r in csv.DictReader(f, delimiter=";"):
                r["_source_naf"] = naf
                r["_source_naf_label"] = naf_label
                rows.append(r)
    seen: dict[str, dict] = {}
    for r in rows:
        s = (r.get("company_siren") or "").strip()
        if not s:
            continue
        cur = seen.get(s)
        if cur is None or float(r.get("overall_score") or 0) > float(cur.get("overall_score") or 0):
            seen[s] = r
    uniq = [r for r in seen.values() if (r.get("is_operating") or "True").lower() != "false"]
    uniq.sort(key=lambda r: float(r.get("overall_score") or 0), reverse=True)
    return uniq[:20]


def clean_company_name(s: str) -> str:
    return s.strip()


def channels_summary(r: dict) -> tuple[str, int]:
    """Return (text, count) — counts pro channels actually usable."""
    items = []
    # Email entreprise OU email perso vérifié (conf>=60 et pas dropée)
    siren = r.get("company_siren")
    if siren in DROP_PERSON_EMAIL:
        pass  # email scrubbed
    else:
        if r.get("person_email") and float(r.get("person_email_conf") or 0) >= 60:
            items.append("email pro nominatif vérifié")
        elif r.get("person_email") and float(r.get("person_email_conf") or 0) >= 30:
            items.append("email pro pattern (à tester)")
    if r.get("person_linkedin") and float(r.get("person_linkedin_conf") or 0) >= 60:
        items.append("LinkedIn perso DRH")
    if r.get("company_phone") and siren not in PHONE_NOTES:
        items.append("tél fixe entreprise")
    if r.get("company_linkedin"):
        items.append("LinkedIn entreprise")
    return (" + ".join(items) or "aucun canal direct", len(items))


# --- Cold email authoring ----------------------------------------------------
# Sector-specific opener angles for GROUPE A NAFs.
SECTOR_ANGLES = {
    "70.22Z": {  # Conseil aux entreprises
        "context": "conseil aux entreprises",
        "pain": "des consultants qui passent 30-40 % de leur temps sur la production de slides, livrables et reporting client",
        "promise": "automatiser la synthèse de comptes-rendus, la rédaction de propositions commerciales et l'analyse documentaire avec un copilote IA encadré (RGPD compris)",
    },
    "78.10Z": {  # Placement
        "context": "recrutement / placement",
        "pain": "le tri manuel de centaines de CV par offre et la rédaction de fiches de poste qui mobilisent vos consultants RH au lieu du contact candidat",
        "promise": "déployer un sourcing IA assisté (matching CV, pré-qualification, rédaction d'annonces personnalisées) en restant 100 % conforme RGPD recrutement",
    },
    "82.99Z": {  # Services aux entreprises
        "context": "services aux entreprises",
        "pain": "des tâches admin répétitives (reporting, mise en forme, traitement de demandes entrantes) qui érodent la marge",
        "promise": "industrialiser ces flux avec des agents IA + automatisations Make/n8n, en commençant par les use-cases les plus chronophages identifiés avec vos équipes",
    },
}


def make_cold_email(r: dict) -> tuple[str, str]:
    """Return (subject, body) ~ 80-120 words."""
    naf = r["_source_naf"]
    angle = SECTOR_ANGLES.get(naf, SECTOR_ANGLES["82.99Z"])
    company = clean_company_name(r["company_name"]).split("(")[0].strip()
    first = (r.get("person_name") or "").split()[0] or "Bonjour"

    subject = f"IA opérationnelle chez {company} — 15 min pour cadrer ?"

    body = (
        f"Bonjour {first},\n\n"
        f"En tant que DRH d'une structure {angle['context']} de 50-99 collaborateurs, "
        f"je devine que vous voyez monter une attente forte côté équipes : pouvoir s'appuyer "
        f"sur l'IA sans bricoler avec des outils grand public.\n\n"
        f"Comeos accompagne depuis 2000 les PME/ETI d'Occitanie sur la montée en compétences. "
        f"Pour {company}, le point sensible que nous voyons souvent : "
        f"{angle['pain']}.\n\n"
        f"Nous avons monté un parcours en 3 niveaux — Découverte (1 j), Approfondissement (2-3 j) "
        f"et Automatisation (5-10 j) — pour {angle['promise']}.\n\n"
        f"Auriez-vous 15 minutes la semaine prochaine pour qu'on regarde ce qui a du sens "
        f"chez vous ? Je peux vous présenter 2-3 cas concrets de notre côté.\n\n"
        f"L'équipe Comeos — Conseil & Formation depuis 2000"
    )
    return subject, body


# --- XLSX writer -------------------------------------------------------------
HEADER = [
    ("Rang", 5),
    ("Entreprise", 40),
    ("SIREN", 12),
    ("NAF", 9),
    ("Secteur (NAF)", 38),
    ("Catégorie GMB", 25),
    ("Taille (Sirene)", 16),
    ("Opérationnel ?", 12),
    ("Ville", 14),
    ("Adresse", 50),
    ("Tél fixe entreprise", 19),
    ("Site web", 38),
    ("LinkedIn entreprise", 50),
    ("Décideur (nom)", 26),
    ("Rôle", 12),
    ("Email pro décideur", 38),
    ("Email confiance", 13),
    ("Note email", 38),
    ("Mobile direct (06/07)", 19),
    ("LinkedIn décideur", 52),
    ("ICP score", 10),
    ("Overall score", 13),
    ("Canaux pro vérifiés", 36),
    ("Cold-email sujet (FR)", 48),
    ("Cold-email corps (FR)", 95),
    ("Flags / qualité", 52),
]


def build_xlsx(rows: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 20 — Comeos IA GROUPE A"

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2C3E50")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col_idx, (label, width) in enumerate(HEADER, start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for idx, r in enumerate(rows, start=1):
        siren = r.get("company_siren")
        subj, body = make_cold_email(r)
        channels_text, channels_n = channels_summary(r)

        # Sanitize email
        p_email = r.get("person_email") or ""
        p_email_conf = r.get("person_email_conf") or ""
        p_email_note = r.get("person_email_note") or ""
        if siren in DROP_PERSON_EMAIL:
            p_email = ""
            p_email_conf = ""
            p_email_note = DROP_PERSON_EMAIL[siren]

        # Sanitize phone label
        company_phone = r.get("company_phone") or ""
        if siren in PHONE_NOTES:
            company_phone = (company_phone + "  ⚠️").strip()

        flags = []
        if siren in DROP_PERSON_EMAIL:
            flags.append("Email DRH retiré (catchall hébergeur)")
        if siren in PHONE_NOTES:
            flags.append(PHONE_NOTES[siren])
        if siren in WEBSITE_NOTES:
            flags.append(WEBSITE_NOTES[siren])
        if p_email and ("PATTERN GUESS" in (r.get("person_email_note") or "")):
            flags.append("Email = pattern guess (NOT verified — test before sending)")
        if not r.get("person_email") and not r.get("person_phone"):
            flags.append("Aucun email/mobile direct → contact via LinkedIn ou standard")
        if not r.get("company_website"):
            flags.append("Pas de site web identifié — vérifier sur mentions légales / Pappers")
        if r.get("company_size") == "21" and not p_email and not r.get("person_phone"):
            pass  # already covered

        flags_text = " | ".join(flags) if flags else "—"

        size_label = {"21": "50-99 emp", "22": "100-199 emp", "31": "200-249 emp"}.get(
            r.get("company_size") or "", r.get("company_size") or ""
        )

        values = [
            idx,
            clean_company_name(r["company_name"]),
            siren,
            r["_source_naf"],
            r["_source_naf_label"],
            r.get("cuisine_type") or "—",
            size_label,
            "Oui" if (r.get("is_operating") or "True").lower() != "false" else "NON",
            r.get("company_city") or "",
            r.get("company_address") or "",
            company_phone or "—",
            r.get("company_website") or "—",
            r.get("company_linkedin") or "—",
            r.get("person_name") or "—",
            r.get("person_role") or "DRH",
            p_email or "—",
            p_email_conf or "—",
            p_email_note or "—",
            r.get("person_phone") or "— (non trouvé via data publique FR)",
            r.get("person_linkedin") or "—",
            r.get("icp_score") or "—",
            r.get("overall_score") or "—",
            channels_text,
            subj,
            body,
            flags_text,
        ]

        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=idx + 1, column=col_idx, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
            # Hyperlinks on URLs + email
            if isinstance(val, str) and (val.startswith("http") or "@" in val and val.count(" ") == 0 and "—" not in val):
                if val.startswith("http"):
                    c.hyperlink = val
                    c.font = Font(color="0000EE", underline="single")
                elif "@" in val and not val.startswith("—"):
                    c.hyperlink = f"mailto:{val}"
                    c.font = Font(color="0000EE", underline="single")

        ws.row_dimensions[idx + 1].height = 220  # leave room for cold email body

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADER))}1"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


def main() -> None:
    rows = load_all_rows()
    print(f"Loaded {len(rows)} unique leads (after dedup + is_operating filter).")
    path = build_xlsx(rows)
    print(f"Wrote {path}")

    # Quick quality recap
    ge50 = sum(1 for r in rows if float(r.get("icp_score") or 0) >= 50)
    have_name = sum(1 for r in rows if r.get("person_name"))
    have_pro_channel = sum(1 for r in rows if channels_summary(r)[1] >= 1)
    have_mobile = sum(1 for r in rows if r.get("person_phone"))
    print(f"  icp_score >= 50 : {ge50}/{len(rows)}")
    print(f"  Décideur nominatif : {have_name}/{len(rows)}")
    print(f"  >=1 canal pro vérifié : {have_pro_channel}/{len(rows)}")
    print(f"  Mobile direct 06/07 : {have_mobile}/{len(rows)}")


if __name__ == "__main__":
    main()
