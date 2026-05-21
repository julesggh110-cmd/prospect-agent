"""Tighten the existing Comeos GROUPE A FINAL deliverable.

Reads output/comeos-ia-services-toulouse-FINAL.xlsx and writes a v2 alongside.

Changes vs v1:
- Rewrites every cold email to fit the 80-120 mots target the brief asked for
  (v1 bodies were 150-160 mots). One template per NAF family:
    * 70.22Z (conseil aux entreprises): productivité doc + livrables consultant
    * 78.10Z (placement/recrutement): tri CV, copilote sourceur, RGPD candidat
    * 82.99Z (autres services support entreprises): back-office, reporting client
- Flags the Roland Gomez duplicate (Lead 13 Winsearch + Lead 15 Akuit
  share the same LinkedIn profile → one is a triangulation error; we keep
  the original entry on the placement-agency match where his profile reads
  as "RH" but tag Akuit with a CROSS-COMPANY note and downgrade its
  verified-channel claim).
"""
from __future__ import annotations

import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parent
SRC = ROOT / "output" / "comeos-ia-services-toulouse-FINAL.xlsx"
DST = ROOT / "output" / "comeos-ia-services-toulouse-FINAL-v2.xlsx"


SIGN_OFF = "L'équipe Comeos — Conseil & Formation depuis 2000"


def _first_name(full: str) -> str:
    if not full:
        return ""
    return full.strip().split()[0]


def _short_company(name: str) -> str:
    if not name:
        return ""
    # Strip parenthetical aliases: "STHREE SAS (HUXLEY - …)" -> "STHREE SAS"
    base = re.split(r"\s*\(", name, maxsplit=1)[0].strip()
    # Cap at ~40 chars to keep the email crisp
    if len(base) > 42:
        base = base[:42].rstrip() + "…"
    return base


def _draft(naf: str, person_name: str, company: str) -> tuple[str, str]:
    """Return (subject, body) tuned per NAF family, body within 80-120 mots."""
    prenom = _first_name(person_name) or "Bonjour"
    co = _short_company(company)

    if naf.startswith("70.22"):
        subject = f"IA pour cabinets de conseil — 15 min avec {co} ?"
        body = (
            f"Bonjour {prenom},\n\n"
            f"En tant que DRH d'un cabinet de conseil de 50-99 collaborateurs, vous voyez "
            f"sûrement vos consultants passer 30-40 % de leur temps sur la production de "
            f"slides, comptes-rendus et propositions commerciales.\n\n"
            f"Comeos (Toulouse, depuis 2000) lance un parcours Formation IA pensé pour "
            f"{co} en 3 niveaux : Découverte (1 j, ChatGPT/Claude productivité), "
            f"Approfondissement (2-3 j, RGPD + cas d'usage métier), Automatisation "
            f"(5-10 j, agents IA branchés sur le CRM).\n\n"
            f"Auriez-vous 15 minutes la semaine prochaine pour qu'on regarde ce qui aurait "
            f"le plus d'impact chez vous ?\n\n"
            f"{SIGN_OFF}"
        )
    elif naf.startswith("78.10") or naf.startswith("78.20") or naf.startswith("78.30"):
        subject = f"Tri CV & sourceur IA — 15 min avec {co} ?"
        body = (
            f"Bonjour {prenom},\n\n"
            f"En tant que DRH d'une agence de placement, vous gérez un flux constant de CV, "
            f"de mises en relation et de reporting client. L'IA bien encadrée peut diviser "
            f"par deux le temps de pré-qualification — sans toucher au RGPD candidat.\n\n"
            f"Comeos (Toulouse, depuis 2000) propose à {co} un parcours Formation IA en 3 "
            f"niveaux : Découverte (1 j), Approfondissement (2-3 j, copilote sourceur + "
            f"matching), Automatisation (5-10 j, agent IA branché ATS).\n\n"
            f"15 minutes la semaine prochaine pour cadrer ce qui a du sens chez vous ?\n\n"
            f"{SIGN_OFF}"
        )
    elif naf.startswith("82.11") or naf.startswith("82.99"):
        subject = f"Back-office IA — 15 min avec {co} ?"
        body = (
            f"Bonjour {prenom},\n\n"
            f"En tant que DRH d'une structure de soutien aux entreprises de 50-99 "
            f"collaborateurs, vous portez à la fois la formation interne et la pression "
            f"d'efficacité côté client. L'IA bien cadrée libère 20-30 % du temps "
            f"back-office (devis, reporting, suivi dossier).\n\n"
            f"Comeos (Toulouse, depuis 2000) propose à {co} un parcours Formation IA en 3 "
            f"niveaux : Découverte (1 j), Approfondissement (2-3 j, RGPD + workflow "
            f"métier), Automatisation (5-10 j, agent IA branché CRM/ERP).\n\n"
            f"15 minutes la semaine prochaine pour échanger ?\n\n"
            f"{SIGN_OFF}"
        )
    else:
        # Fallback (shouldn't trigger on this dataset, all are A-group)
        subject = f"Formation IA pour vos équipes — 15 min avec {co} ?"
        body = (
            f"Bonjour {prenom},\n\n"
            f"En tant que DRH d'une PME de 50-99 collaborateurs, vous êtes en première "
            f"ligne sur la montée en compétences IA des équipes — un sujet où ChatGPT "
            f"grand public ne suffit plus dès qu'on parle données client.\n\n"
            f"Comeos (Toulouse, depuis 2000) propose à {co} un parcours Formation IA en 3 "
            f"niveaux : Découverte (1 j), Approfondissement (2-3 j, RGPD), Automatisation "
            f"(5-10 j, agents custom branchés sur vos outils).\n\n"
            f"15 minutes la semaine prochaine pour cadrer ?\n\n"
            f"{SIGN_OFF}"
        )

    return subject, body


def _word_count(text: str) -> int:
    return len(re.findall(r"\S+", text))


def main() -> None:
    wb = load_workbook(SRC)
    ws = wb.worksheets[0]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    naf_c = col["NAF"]
    ent_c = col["Entreprise"]
    dec_c = col["Décideur (nom)"]
    sujet_c = col["Cold-email sujet (FR)"]
    corps_c = col["Cold-email corps (FR)"]
    flags_c = col["Flags / qualité"]
    channels_c = col["Canaux pro vérifiés"]
    overall_c = col["Overall score"]
    li_pers_c = col["LinkedIn décideur"]
    email_c = col["Email pro décideur"]
    email_conf_c = col["Email confiance"]
    email_note_c = col["Note email"]

    # --- pass 1: rewrite cold emails ---
    word_counts: list[int] = []
    for row_i in range(2, ws.max_row + 1):
        naf = (ws.cell(row=row_i, column=naf_c).value or "").strip()
        company = (ws.cell(row=row_i, column=ent_c).value or "").strip()
        person = (ws.cell(row=row_i, column=dec_c).value or "").strip()
        sujet, corps = _draft(naf, person, company)
        ws.cell(row=row_i, column=sujet_c, value=sujet)
        ws.cell(row=row_i, column=corps_c, value=corps).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        word_counts.append(_word_count(corps))

    # --- pass 2: cross-company duplicate flag (Roland Gomez at Lead 13 + 15) ---
    # Detect any decideur LinkedIn URL that appears on more than one company.
    li_to_rows: dict[str, list[int]] = {}
    for row_i in range(2, ws.max_row + 1):
        li = (ws.cell(row=row_i, column=li_pers_c).value or "").strip().lower()
        if not li:
            continue
        # Normalise: strip trailing slash + protocol
        li = re.sub(r"^https?://", "", li).rstrip("/")
        li_to_rows.setdefault(li, []).append(row_i)

    cross_rows: set[int] = set()
    for li, rows in li_to_rows.items():
        if len(rows) > 1:
            # Keep the highest-overall-score row, flag the others
            rows_sorted = sorted(
                rows,
                key=lambda r: (ws.cell(row=r, column=overall_c).value or 0),
                reverse=True,
            )
            for r in rows_sorted[1:]:
                cross_rows.add(r)

    for row_i in cross_rows:
        existing_flags = ws.cell(row=row_i, column=flags_c).value or ""
        note = (
            "CROSS-COMPANY: LinkedIn décideur déjà attribué à une autre société du fichier "
            "— vérifier manuellement avant outreach"
        )
        new_flags = (
            note if existing_flags in ("", "—") else f"{note} | {existing_flags}"
        )
        ws.cell(row=row_i, column=flags_c, value=new_flags)

        # Downgrade verified channels: LinkedIn perso n'est plus fiable
        ch = (ws.cell(row=row_i, column=channels_c).value or "").strip()
        ch = ch.replace("LinkedIn perso DRH", "LinkedIn perso (à reconfirmer)")
        ws.cell(row=row_i, column=channels_c, value=ch)

        # Knock 6 points off the overall score (manual confidence haircut)
        sc = ws.cell(row=row_i, column=overall_c).value
        if isinstance(sc, (int, float)):
            ws.cell(row=row_i, column=overall_c, value=max(0, int(sc) - 6))

    # --- pass 3: persist & report ---
    wb.save(DST)

    print(f"Saved: {DST}")
    print(f"Cold-email word counts: min={min(word_counts)} max={max(word_counts)} avg={sum(word_counts)//len(word_counts)}")
    print(f"Cross-company rows downgraded: {sorted(cross_rows)}")
    in_range = sum(1 for n in word_counts if 80 <= n <= 120)
    print(f"Within 80-120 mots: {in_range}/{len(word_counts)}")


if __name__ == "__main__":
    main()
