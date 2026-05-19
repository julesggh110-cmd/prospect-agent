"""
Sheets export — push leads to Google Sheets, with Excel-FR-friendly CSV fallback.

CSV format choice matters:
- Excel FR opens comma-CSV as a single column (it expects `;`).
- Plain `,` CSVs with internal quotes can break older parsers.
- A UTF-8 BOM + `;` delimiter opens cleanly in Excel FR, Excel EN, Numbers,
  LibreOffice, and Google Sheets ("Import" → auto-detects).

Two auth modes for Sheets:
1. Service account JSON file (path in env var GOOGLE_SERVICE_ACCOUNT_JSON)
   → simplest for automation. Create a Cloud project, enable the Sheets API,
     download a service account JSON, share the target sheet (edit) with the
     service-account email.
2. (TBD) OAuth user flow — not implemented yet.

If Sheets isn't configured, we write **two** files side by side:
- `leads-<timestamp>.csv` (semicolon, UTF-8 BOM) — opens in Excel
- `leads-<timestamp>.xlsx` (real Excel binary) — opens in everything
"""
from __future__ import annotations

import csv
import os
import time
from pathlib import Path
from typing import Optional

try:
    import gspread  # type: ignore
    from google.oauth2.service_account import Credentials  # type: ignore
except ImportError:  # pragma: no cover
    gspread = None  # type: ignore
    Credentials = None  # type: ignore

# Imported lazily to keep this module light if Lead model isn't loaded
HEADERS = [
    "icp_score",
    "company_name", "company_siren", "company_naf_label",
    "company_city", "company_address", "company_size", "company_website",
    "company_email",      # generic shared inbox (contact@, info@) — NOT the person
    "company_phone", "company_phone_conf",
    "company_linkedin", "company_linkedin_conf",
    "company_instagram", "company_instagram_conf",
    "company_facebook",
    "person_name", "person_name_conf",
    "person_role", "person_role_conf",
    "person_email", "person_email_conf", "person_email_note",
    "person_phone", "person_phone_conf",
    "person_linkedin", "person_linkedin_conf",
    "person_instagram", "person_instagram_conf",
    "overall_score",
    "is_new_lead",
    # Cold email (only filled when --generate-emails was passed)
    "cold_email_subject", "cold_email_body", "cold_email_angle",
    "dropped", "drop_reason",
]


def _row_for(lead) -> list:  # `lead` is a triangulation.Lead but we keep this lazy
    def scored(field):
        return [field.value or "", field.confidence if field.value else ""]

    return [
        getattr(lead, "icp_score", "") or "",
        lead.company_name,
        lead.company_siren or "",
        lead.company_naf_label or "",
        lead.company_city or "",
        lead.company_address or "",
        lead.company_size or "",
        lead.company_website or "",
        lead.company_email or "",
        *scored(lead.company_phone),
        *scored(lead.company_linkedin),
        *scored(lead.company_instagram),
        lead.company_facebook or "",
        *scored(lead.person_name),
        *scored(lead.person_role),
        lead.person_email.value or "",
        lead.person_email.confidence if lead.person_email.value else "",
        lead.person_email.note or "",
        *scored(lead.person_phone),
        *scored(lead.person_linkedin),
        *scored(lead.person_instagram),
        lead.overall_score,
        "new" if getattr(lead, "is_new_lead", False) else "",
        getattr(lead, "cold_email_subject", "") or "",
        getattr(lead, "cold_email_body", "") or "",
        getattr(lead, "cold_email_angle", "") or "",
        "yes" if lead.dropped else "",
        lead.drop_reason or "",
    ]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_leads(leads: list, *, prefer_sheet: bool = True) -> str:
    """Push leads to Google Sheets if configured, else CSV. Returns URL or path."""
    if prefer_sheet:
        sheet_id = os.environ.get("DEFAULT_SHEET_ID", "").strip()
        sa_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
        if sheet_id and sa_path and gspread is not None:
            try:
                return _push_to_sheet(leads, sheet_id, sa_path)
            except Exception as exc:
                print(f"[sheets] Push failed ({exc}); falling back to CSV.")
    return _write_csv(leads)


def _push_to_sheet(leads: list, sheet_id: str, sa_path: str) -> str:
    if gspread is None or Credentials is None:
        raise RuntimeError("gspread / google-auth not installed")
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(sa_path, scopes=scopes)
    client = gspread.authorize(creds)
    sh = client.open_by_key(sheet_id)
    ws_name = time.strftime("leads-%Y-%m-%d-%H%M%S")
    ws = sh.add_worksheet(title=ws_name, rows=max(2, len(leads) + 5), cols=len(HEADERS))
    rows = [HEADERS] + [_row_for(l) for l in leads]
    ws.update(values=rows, range_name="A1")
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit#gid={ws.id}"


def _write_premium_xlsx(rows: list[list], xlsx_path: Path) -> None:
    """Write the rows to an XLSX with rich formatting.

    Visuals:
    - Header row: bold white text on dark teal background, frozen.
    - Confidence columns (*_conf): color scale red->yellow->green based on value.
    - URL columns (website / linkedin / instagram / facebook): clickable hyperlinks.
    - Column widths auto-sized (capped at 60 chars).
    - Filter enabled on header.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()
    ws = wb.active
    ws.title = "leads"

    header = rows[0]
    data_rows = rows[1:]

    # Identify special column groups
    conf_cols = {i + 1 for i, h in enumerate(header) if isinstance(h, str) and h.endswith("_conf")}
    url_cols = {
        i + 1 for i, h in enumerate(header)
        if isinstance(h, str) and any(k in h for k in ("website", "linkedin", "instagram", "facebook"))
        and not h.endswith("_conf")
    }

    # Header styling
    header_fill = PatternFill("solid", fgColor="14b8a6")  # teal
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, value in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Data rows
    for r_idx, row in enumerate(data_rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            v = "" if value is None else value
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            if c_idx in url_cols and isinstance(v, str) and v.startswith(("http://", "https://")):
                cell.hyperlink = v
                cell.font = Font(color="0563C1", underline="single")

    # Freeze + filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Color scale on confidence columns (red <40, yellow 60, green >80)
    if data_rows:
        last_row = len(data_rows) + 1
        rule = ColorScaleRule(
            start_type="num", start_value=0,  start_color="FCA5A5",  # red-300
            mid_type="num",   mid_value=60,  mid_color="FDE68A",     # yellow-200
            end_type="num",   end_value=100, end_color="86EFAC",     # green-300
        )
        for c_idx in conf_cols:
            col_letter = get_column_letter(c_idx)
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{last_row}", rule)

    # Auto column widths (cap 60)
    for col_idx in range(1, len(header) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

    # Row height for header
    ws.row_dimensions[1].height = 22

    wb.save(xlsx_path)


def _write_csv(leads: list, out_dir: Optional[Path] = None) -> str:
    """Write two artifacts: Excel-FR-friendly CSV (semicolon + BOM) AND XLSX.

    Returns the CSV path (the XLSX sits alongside with the same stem).
    """
    out_dir = out_dir or Path(__file__).resolve().parent / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = time.strftime("leads-%Y-%m-%d-%H%M%S")
    csv_path = out_dir / f"{stem}.csv"
    xlsx_path = out_dir / f"{stem}.xlsx"

    rows = [HEADERS] + [_row_for(l) for l in leads]

    # CSV with `;` + UTF-8 BOM (Excel FR opens it as a real table, no Import wizard)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)

    # Premium XLSX: bold header w/ background, clickable hyperlinks on URL
    # cells, color-graded confidence scores, auto-column widths, frozen header.
    try:
        _write_premium_xlsx(rows, xlsx_path)
    except ImportError:
        pass  # openpyxl missing → CSV-only is still fine

    return str(csv_path)
